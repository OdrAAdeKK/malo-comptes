from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime, date
import os
from models import Musicien, Operation, Concert, Participation, Report, db
from sqlalchemy import func
from mes_utils import calculer_credit_actuel
from collections import defaultdict
import locale

# Liste manuelle des mois avec accents, pour compatibilité Windows
MOIS_FR = [
    "janvier", "février", "mars", "avril", "mai", "juin",
    "juillet", "août", "septembre", "octobre", "novembre", "décembre"
]
def mois_francais(dt):
    try:
        mois = MOIS_FR[dt.month - 1]
        return f"{mois} {dt.year}"
    except Exception:
        return dt.strftime('%m/%Y')

# Tentative de locale (pour Linux/Mac)
try:
    locale.setlocale(locale.LC_TIME, 'fr_FR.UTF-8')
except:
    try:
        locale.setlocale(locale.LC_TIME, 'fr_FR')
    except:
        pass

def generer_export_excel():
    aujourd_hui = date.today()
    wb = Workbook()
    ws = wb.active
    ws.title = "Comptes"

    musiciens = Musicien.query.filter(Musicien.actif == True, Musicien.type != 'structure').all()
    musiciens.sort(key=lambda m: (m.nom != "ARNOULD" or m.prenom != "Jérôme", m.nom, m.prenom))

    structures = Musicien.query.filter(
        Musicien.type == 'structure',
        ~Musicien.nom.in_(["CB ASSO7", "CAISSE ASSO7", "TRESO ASSO7"])
    ).all()
    cb = Musicien.query.filter_by(nom="CB ASSO7").first()
    caisse = Musicien.query.filter_by(nom="CAISSE ASSO7").first()
    tous = [*musiciens, *structures, cb, caisse]
    tous_avec_treso = tous + [Musicien(prenom="", nom="TRESO ASSO7")]

    pastel_colors = [
        "FFEBEE", "E3F2FD", "E8F5E9", "FFFDE7", "F3E5F5",
        "FBE9E7", "D2F8E5", "FFF1E0", "CDEDF6", "D8C9FF"
    ]

    # Ligne 1 : noms fusionnés par 6 colonnes, gras et taille 16
    ws.cell(row=1, column=1, value="Musicien/Structure").font = Font(bold=True)
    for idx, m in enumerate(tous_avec_treso):
        col_start = 2 + idx * 6
        ws.merge_cells(start_row=1, start_column=col_start, end_row=1, end_column=col_start + 5)
        nom_complet = f"{m.prenom + ' ' if m.prenom else ''}{m.nom}"
        cell = ws.cell(row=1, column=col_start, value=nom_complet)
        cell.alignment = Alignment(horizontal="center")
        cell.font = Font(bold=True, size=16)
        fill = PatternFill(start_color=pastel_colors[idx % len(pastel_colors)],
                           end_color=pastel_colors[idx % len(pastel_colors)], fill_type="solid")
        for c in range(col_start, col_start + 6):
            ws.cell(row=1, column=c).fill = fill

    # Rubriques fixes
    ws.cell(row=3, column=1, value="CRÉDIT ACTUEL").font = Font(bold=True)
    ws.cell(row=4, column=1, value="GAINS À VENIR").font = Font(bold=True)
    ws.cell(row=5, column=1, value="CRÉDIT POTENTIEL").font = Font(bold=True)
    ws.cell(row=7, column=1, value="REPORTS").font = Font(bold=True)

    concerts = Concert.query.all()
    valeurs = {}

    for idx, m in enumerate(tous_avec_treso):
        nom = m.nom
        col = 2 + idx * 6
        fill = PatternFill(start_color=pastel_colors[idx % len(pastel_colors)],
                           end_color=pastel_colors[idx % len(pastel_colors)], fill_type="solid")
        if nom == "TRESO ASSO7":
            cb_vals = valeurs.get("CB ASSO7", {})
            caisse_vals = valeurs.get("CAISSE ASSO7", {})
            credit = (cb_vals.get("credit", 0) + caisse_vals.get("credit", 0))
            report = (cb_vals.get("report", 0) + caisse_vals.get("report", 0))
            gains_a_venir = (cb_vals.get("gains_a_venir", 0) + caisse_vals.get("gains_a_venir", 0))
        else:
            credit = calculer_credit_actuel(m, concerts)
            report = db.session.query(func.sum(Report.montant)).filter_by(musicien_id=m.id).scalar() or 0
            credit += report
            ops_avenir = db.session.query(Operation).filter(Operation.musicien_id == m.id, Operation.date > aujourd_hui).all()
            total_ops_avenir = sum(op.montant if op.type == 'crédit' else -op.montant for op in ops_avenir)
            part_pot = db.session.query(func.sum(Participation.credit_calcule_potentiel))\
                .filter_by(musicien_id=m.id).scalar() or 0
            gains_a_venir = part_pot + total_ops_avenir
            if m.nom not in ["ASSO7"]:
                concerts_cible = Concert.query.filter(
                    Concert.paye == False,
                    Concert.mode_paiement_prevu.ilike(f"%{m.nom}%")
                ).all()
                for concert in concerts_cible:
                    gains_a_venir += concert.recette_attendue or 0
            valeurs[nom] = {
                "credit": round(credit, 2),
                "report": round(report, 2),
                "gains_a_venir": round(gains_a_venir, 2)
            }
        font_14_bold = Font(bold=True, size=14)
        ws.cell(row=3, column=col + 4, value=round(credit, 2)).fill = fill
        ws.cell(row=4, column=col + 4, value=round(gains_a_venir, 2)).fill = fill
        ws.cell(row=5, column=col + 4, value=round(credit + gains_a_venir, 2)).fill = fill
        ws.cell(row=7, column=col + 4, value=round(report, 2)).fill = fill
        for row in [2, 6, 8]:
            for i in range(6):
                ws.cell(row=row, column=col + i).fill = fill
        for i, champ in enumerate(["Date", "Type", "Motif", "Détail", "Montant", "Concert"]):
            cell = ws.cell(row=9, column=col + i, value=champ)
            cell.font = Font(bold=True)
            cell.fill = fill

    mouvements_par_personne = defaultdict(list)
    for idx, m in enumerate(tous_avec_treso):
        nom = m.nom
        if nom == "TRESO ASSO7":
            continue
        operations = Operation.query.filter_by(musicien_id=m.id).all()
        for op in operations:
            label = ""
            if op.concert:
                c = op.concert
                label = f"{c.lieu} - {c.date.strftime('%d/%m/%Y')}"
            mouvements_par_personne[nom].append({
                "date": op.date,
                "type": "crédit" if op.type == "credit" else "débit",
                "motif": op.motif,
                "detail": op.precision,
                "montant": op.montant if op.type == "credit" else -op.montant,
                "concert": label
            })
        participations = Participation.query.filter_by(musicien_id=m.id).all()
        for p in participations:
            c = p.concert
            if c:
                montant = p.credit_calcule if c.paye else p.credit_calcule_potentiel
                label = f"{c.lieu} - {c.date.strftime('%d/%m/%Y')}"
                mouvements_par_personne[nom].append({
                    "date": c.date,
                    "type": "crédit",
                    "motif": "participation",
                    "detail": "",
                    "montant": montant,
                    "concert": label
                })
        if m.nom in ["CB ASSO7", "CAISSE ASSO7"]:
            concerts_recettes = Concert.query.filter(
                Concert.paye == False,
                Concert.mode_paiement_prevu == m.nom
            ).all()
            for c in concerts_recettes:
                label = f"{c.lieu} - {c.date.strftime('%d/%m/%Y')}"
                mouvements_par_personne[m.nom].append({
                    "date": c.date,
                    "type": "crédit",
                    "motif": "Recette attendue",
                    "detail": "",
                    "montant": c.recette_attendue or 0,
                    "concert": label
                })

    row_base = 10
    max_rows = row_base
    for idx, m in enumerate(tous_avec_treso):
        nom = m.nom
        if nom == "TRESO ASSO7":
            continue
        col = 2 + idx * 6
        mouvements = mouvements_par_personne[nom]
        mouvements.sort(key=lambda x: x['date'] or aujourd_hui, reverse=True)
        last_month = None
        row = row_base
        fill = PatternFill(start_color=pastel_colors[idx % len(pastel_colors)],
                           end_color=pastel_colors[idx % len(pastel_colors)], fill_type="solid")
        for mouv in mouvements:
            if mouv['date']:
                mois_annee = mois_francais(mouv['date'])
                if mois_annee != last_month:
                    ws.cell(row=row, column=col, value=mois_annee).font = Font(bold=True)
                    for i in range(6):
                        ws.cell(row=row, column=col + i).fill = fill
                    row += 1
                    last_month = mois_annee
            ws.cell(row=row, column=col + 0, value=mouv['date'].strftime("%d/%m/%Y") if mouv['date'] else "")
            ws.cell(row=row, column=col + 1, value=mouv['type'])
            ws.cell(row=row, column=col + 2, value=mouv['motif'])
            ws.cell(row=row, column=col + 3, value=mouv['detail'])
            ws.cell(row=row, column=col + 4, value=round(mouv['montant'], 2))
            ws.cell(row=row, column=col + 5, value=mouv['concert'])
            for i in range(6):
                ws.cell(row=row, column=col + i).fill = fill
            row += 1
        max_rows = max(max_rows, row)

    # Ajustement automatique des largeurs de colonnes
    for col in range(1, ws.max_column + 1):
        max_length = 0
        for row in range(1, max_rows + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col)].width = max_length + 2

    dossier = "exports"
    os.makedirs(dossier, exist_ok=True)
    chemin = os.path.join(dossier, f"Comptes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    wb.save(chemin)
    return chemin
